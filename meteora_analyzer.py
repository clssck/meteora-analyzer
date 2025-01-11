from __future__ import annotations

import argparse
import asyncio
import csv
import logging
import logging.handlers
import sys
import time
from contextlib import suppress
from dataclasses import dataclass
from pathlib import Path
from typing import TYPE_CHECKING, Any, ClassVar

import aiohttp
import openpyxl
from openpyxl.styles import Font, PatternFill

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet


# Configuration
class Config:
    # API endpoints
    METEORA_API: ClassVar[str] = "https://dlmm-api.meteora.ag"
    METEORA_API_PAIR_ENDPOINT: ClassVar[str] = "/pair/all"
    DEX_SCREENER_API_URL: ClassVar[str] = (
        "https://api.dexscreener.com/latest/dex/pairs/solana"
    )
    JUPITER_TOKEN_STRICT_LIST_API: ClassVar[str] = "https://token.jup.ag/strict"
    JUPITER_TOKEN_ALL_LIST_API: ClassVar[str] = "https://token.jup.ag/all"

    # Rate limiting
    REQUESTS_PER_MINUTE: ClassVar[int] = 300  # DexScreener limit
    BASE_DELAY: ClassVar[float] = 60 / REQUESTS_PER_MINUTE
    MAX_RETRIES: ClassVar[int] = 5
    CONCURRENT_LIMIT: ClassVar[int] = 20  # Process 600 pairs per batch
    CHUNK_SIZE: ClassVar[int] = 30  # DexScreener max addresses per request
    REQUEST_TIMEOUT: ClassVar[int] = 30

    # Safety buffer for rate limits (use 80% of max rate)
    RATE_LIMIT_BUFFER: ClassVar[float] = 0.8
    EFFECTIVE_RATE_LIMIT: ClassVar[int] = int(REQUESTS_PER_MINUTE * RATE_LIMIT_BUFFER)

    # Filtering thresholds
    MIN_LIQUIDITY_THRESHOLD: ClassVar[float] = 1000  # Minimum $1k TVL
    MIN_BIN_STEP: ClassVar[int] = 1
    MAX_BIN_STEP: ClassVar[int] = 400  # Skip extreme bin steps
    MIN_BASE_FEE: ClassVar[float] = 0.01  # 0.01% minimum fee
    MAX_BASE_FEE: ClassVar[float] = 10.0  # 10% maximum fee

    # Scoring thresholds
    VOLUME_THRESHOLDS: ClassVar[dict[str, int]] = {
        "HIGH": 1_000_000,
        "MEDIUM": 100_000,
        "LOW": 10_000,
    }
    TVL_THRESHOLDS: ClassVar[dict[str, int]] = {
        "HIGH": 500_000,
        "MEDIUM": 50_000,
        "LOW": 5_000,
    }

    # Logging
    LOG_FILE: ClassVar[str] = "meteora_analyzer.log"
    LOG_LEVEL: ClassVar[int] = logging.INFO

    # Blue chip tokens
    BLUE_CHIPS: ClassVar[list[str]] = [
        token.lower()
        for token in [
            "USDC",
            "SOL",
            "USDT",
            "jitoSOL",
            "bSOL",
            "JupSOL",
            "INF",
            "JLP",
            "WBTC",
            "WETH",
            "bonkSOL",
            "LST",
            "mSOL",
            "zippySOL",
        ]
    ]

    # Time constants
    SECONDS_IN_MINUTE: ClassVar[int] = 60

    # Scoring thresholds
    VOLATILITY_THRESHOLD: ClassVar[float] = 1.2
    EXTREME_VOLATILITY_THRESHOLD: ClassVar[float] = 3.0
    LOW_CAPITAL_EFFICIENCY_THRESHOLD: ClassVar[float] = 0.5
    LOW_FEE_YIELD_THRESHOLD: ClassVar[float] = 0.3

    # Bin step ranges
    IDEAL_BIN_STEP_MIN: ClassVar[int] = 20
    IDEAL_BIN_STEP_MAX: ClassVar[int] = 100
    ACCEPTABLE_BIN_STEP_MIN: ClassVar[int] = 10
    ACCEPTABLE_BIN_STEP_MAX: ClassVar[int] = 200
    EXTREME_BIN_STEP_MAX: ClassVar[int] = 300

    # Investment score thresholds
    STRONG_BUY_SCORE: ClassVar[int] = 75
    BUY_SCORE: ClassVar[int] = 55
    HOLD_SCORE: ClassVar[int] = 35

    # Risk rating thresholds
    LOW_RISK_RATING: ClassVar[int] = 2
    MEDIUM_RISK_RATING: ClassVar[int] = 3
    HIGH_RISK_RATING: ClassVar[int] = 4


@dataclass
class ChunkResults:
    pairs: list[dict[str, Any]]
    processed: int
    errors: int


class RateLimiter:
    def __init__(self, max_rate: float):
        self.max_rate = max_rate
        self.last_request_time = 0.0
        self.request_times = []  # Track last minute of requests

    async def wait(self):
        current_time = time.time()

        # Remove requests older than 1 minute
        self.request_times = [
            t for t in self.request_times if current_time - t < Config.SECONDS_IN_MINUTE
        ]

        # If we're at the rate limit, wait until we have capacity
        if len(self.request_times) >= Config.EFFECTIVE_RATE_LIMIT:
            wait_time = 60 - (current_time - self.request_times[0])
            if wait_time > 0:
                await asyncio.sleep(wait_time)

        # Add current request
        self.request_times.append(current_time)

        # Small delay between requests to prevent bursts
        elapsed = current_time - self.last_request_time
        if elapsed < Config.BASE_DELAY:
            await asyncio.sleep(Config.BASE_DELAY - elapsed)

        self.last_request_time = time.time()


class ApiClient:
    HTTP_TOO_MANY_REQUESTS = 429

    def __init__(self):
        self.rate_limiter = RateLimiter(Config.REQUESTS_PER_MINUTE / 60)
        self.session = aiohttp.ClientSession(
            timeout=aiohttp.ClientTimeout(total=Config.REQUEST_TIMEOUT),
            headers={
                "Accept": "application/json",
                "User-Agent": "Meteora Analyzer Bot",
            },
        )

    async def fetch(self, url: str) -> dict | None:
        await self.rate_limiter.wait()
        try:
            async with self.session.get(url) as response:
                if response.status == self.HTTP_TOO_MANY_REQUESTS:
                    await asyncio.sleep(Config.BASE_DELAY * 2)
                    return None
                response.raise_for_status()
                return await response.json()
        except Exception:
            logging.exception("Failed to fetch %s", url)
            return None


class DataProcessor:
    @staticmethod
    def validate_pair_data(pair: dict) -> bool:
        required_fields = [
            "pairAddress",
            "baseToken",
            "quoteToken",
            "volume",
            "liquidity",
        ]
        return all(field in pair for field in required_fields)

    @staticmethod
    def enrich_pair_data(
        pair: dict,
        token_map: dict[str, dict[str, Any]],
        meteora_data: list[dict[str, Any]],
    ) -> dict | None:
        if not DataProcessor.validate_pair_data(pair):
            return None

        pair_address = pair["pairAddress"]
        meteora_pair = next(
            (p for p in meteora_data if p["address"] == pair_address),
            None,
        )
        if not meteora_pair:
            return None

        # Enrich token data
        base_token = pair["baseToken"]
        quote_token = pair["quoteToken"]
        pair["baseToken"] = {**base_token, **token_map.get(base_token["address"], {})}
        pair["quoteToken"] = {
            **quote_token,
            **token_map.get(quote_token["address"], {}),
        }

        # Add additional metrics
        pair["strict"] = token_map.get(base_token["address"], {}).get(
            "strict",
            False,
        ) and token_map.get(quote_token["address"], {}).get("strict", False)
        pair["bluechip"] = (
            base_token["symbol"].lower() in Config.BLUE_CHIPS
            and quote_token["symbol"].lower() in Config.BLUE_CHIPS
        )

        # Add Meteora data
        pair["bin_step"] = meteora_pair.get("bin_step")
        pair["base_fee"] = float(meteora_pair.get("base_fee_percentage", 0)) / 100

        # Add TVL and FDV
        pair["tvl"] = pair["liquidity"]["usd"]
        pair["fdv"] = pair.get("fdv", 0)

        # Calculate volume projections for different timeframes
        pair["volume"] = {
            "raw": {  # Raw volumes from API
                "m5": pair["volume"]["m5"],
                "h1": pair["volume"]["h1"],
                "h6": pair["volume"]["h6"],
                "h24": pair["volume"]["h24"],
            },
            "projected": {  # Projected 24h volumes based on each timeframe
                "m5": pair["volume"]["m5"] * 288,  # 5min -> 24h (288 periods)
                "h1": pair["volume"]["h1"] * 24,  # 1h -> 24h (24 periods)
                "h6": pair["volume"]["h6"] * 4,  # 6h -> 24h (4 periods)
                "h24": pair["volume"]["h24"],  # actual 24h volume
            },
        }

        # Calculate fees for each timeframe
        pair["fees"] = {
            "raw": {  # Raw fees from each period
                "m5": pair["base_fee"] * pair["volume"]["raw"]["m5"],
                "h1": pair["base_fee"] * pair["volume"]["raw"]["h1"],
                "h6": pair["base_fee"] * pair["volume"]["raw"]["h6"],
                "h24": pair["base_fee"] * pair["volume"]["raw"]["h24"],
            },
            "projected": {  # Projected 24h fees based on each timeframe
                "m5": pair["base_fee"] * pair["volume"]["projected"]["m5"],
                "h1": pair["base_fee"] * pair["volume"]["projected"]["h1"],
                "h6": pair["base_fee"] * pair["volume"]["projected"]["h6"],
                "h24": pair["base_fee"] * pair["volume"]["projected"]["h24"],
            },
        }

        # Calculate volume and fee ratios
        pair["ratios"] = {
            "volume": {
                "5m_to_1h": pair["volume"]["raw"]["m5"]
                / pair["volume"]["raw"]["h1"]
                * 12
                if pair["volume"]["raw"]["h1"] > 0
                else 0,
                "1h_to_6h": pair["volume"]["raw"]["h1"]
                / pair["volume"]["raw"]["h6"]
                * 6
                if pair["volume"]["raw"]["h6"] > 0
                else 0,
                "6h_to_24h": pair["volume"]["raw"]["h6"]
                / pair["volume"]["raw"]["h24"]
                * 4
                if pair["volume"]["raw"]["h24"] > 0
                else 0,
            },
            "volume_to_tvl": {
                "m5": pair["volume"]["raw"]["m5"] / pair["tvl"]
                if pair["tvl"] > 0
                else 0,
                "h1": pair["volume"]["raw"]["h1"] / pair["tvl"]
                if pair["tvl"] > 0
                else 0,
                "h6": pair["volume"]["raw"]["h6"] / pair["tvl"]
                if pair["tvl"] > 0
                else 0,
                "h24": pair["volume"]["raw"]["h24"] / pair["tvl"]
                if pair["tvl"] > 0
                else 0,
            },
            "fees_to_tvl": {
                "m5": pair["fees"]["raw"]["m5"] / pair["tvl"] * 100
                if pair["tvl"] > 0
                else 0,
                "h1": pair["fees"]["raw"]["h1"] / pair["tvl"] * 100
                if pair["tvl"] > 0
                else 0,
                "h6": pair["fees"]["raw"]["h6"] / pair["tvl"] * 100
                if pair["tvl"] > 0
                else 0,
                "h24": pair["fees"]["raw"]["h24"] / pair["tvl"] * 100
                if pair["tvl"] > 0
                else 0,
            },
        }

        # Calculate volume momentum (>1 means accelerating, <1 means decelerating)
        pair["momentum"] = {
            "5m": pair["ratios"]["volume"]["5m_to_1h"],
            "1h": pair["ratios"]["volume"]["1h_to_6h"],
            "6h": pair["ratios"]["volume"]["6h_to_24h"],
        }

        # Calculate opportunity score (0-100)
        score_components = {
            # Volume is still important but we care more about volume/TVL ratio
            "capital_efficiency": min(pair["ratios"]["volume_to_tvl"]["h24"], 3)
            * 25,  # 25 points for good capital utilization
            # Dynamic fees are a key DLMM feature
            "fee_yield": min(pair["ratios"]["fees_to_tvl"]["h24"] / 1, 1)
            * 30,  # 30 points for good fee yield
            # Volatility is good for DLMM (dynamic fees increase)
            "volatility": (
                (pair["momentum"]["5m"] > Config.VOLATILITY_THRESHOLD)
                * 5  # Reward higher volatility
                + (pair["momentum"]["1h"] > Config.VOLATILITY_THRESHOLD) * 5
                + (pair["momentum"]["6h"] > Config.VOLATILITY_THRESHOLD) * 5
            ),
            # Minimum liquidity still matters but less important
            "base_liquidity": min(pair["tvl"] / 10_000, 1)
            * 10,  # 10 points for basic liquidity
            # Bin step is important for concentrated liquidity
            "bin_quality": (
                10
                if (
                    pair["bin_step"] is not None
                    and Config.IDEAL_BIN_STEP_MIN
                    <= int(pair["bin_step"])
                    <= Config.IDEAL_BIN_STEP_MAX
                )
                # Ideal range
                else 5
                if (
                    pair["bin_step"] is not None
                    and Config.ACCEPTABLE_BIN_STEP_MIN
                    <= int(pair["bin_step"])
                    <= Config.ACCEPTABLE_BIN_STEP_MAX
                )
                # Acceptable range
                else 0  # Too narrow or wide
            ),
            # Safety is still relevant but less weighted
            "safety": (
                (8 if pair["strict"] else 0)  # 8 points for strict list
                + (7 if pair["bluechip"] else 0)  # 7 points for bluechip
            ),
        }

        pair["scores"] = {
            "components": score_components,
            "total": sum(score_components.values()),
        }

        # Calculate risk rating (1-5, 1 being safest)
        risk_factors = {
            "low_capital_efficiency": pair["ratios"]["volume_to_tvl"]["h24"]
            < Config.LOW_CAPITAL_EFFICIENCY_THRESHOLD,  # Low capital usage
            "low_fee_yield": pair["ratios"]["fees_to_tvl"]["h24"]
            < Config.LOW_FEE_YIELD_THRESHOLD,  # Low fee generation
            "extreme_volatility": any(
                m > Config.EXTREME_VOLATILITY_THRESHOLD
                for m in pair["momentum"].values()
            ),  # Too volatile
            "poor_bin_step": (
                pair["bin_step"] is None
                or int(pair["bin_step"]) < Config.ACCEPTABLE_BIN_STEP_MIN
                or int(pair["bin_step"]) > Config.EXTREME_BIN_STEP_MAX
            ),  # Poor bin configuration
            "non_strict": not pair["strict"],  # Not on Jupiter strict list
            "non_bluechip": not pair["bluechip"],  # Not a bluechip pair
        }

        risk_score = sum(risk_factors.values())
        pair["risk"] = {
            "factors": risk_factors,
            "rating": min(max(risk_score + 1, 1), 5),  # Convert to 1-5 scale
            "description": {
                1: "Very Safe - Efficient capital use with good fee generation",
                2: "Safe - Good metrics with manageable volatility",
                3: "Moderate - Decent opportunity but watch bin positioning",
                4: "High - Capital efficiency or fee concerns",
                5: "Very High - Multiple efficiency/safety concerns",
            }[min(max(risk_score + 1, 1), 5)],
        }

        # Add investment recommendation
        if (
            pair["scores"]["total"] >= Config.STRONG_BUY_SCORE
            and pair["risk"]["rating"] <= Config.LOW_RISK_RATING
        ):
            recommendation = "Strong Buy - Well positioned DLMM"
        elif (
            pair["scores"]["total"] >= Config.BUY_SCORE
            and pair["risk"]["rating"] <= Config.MEDIUM_RISK_RATING
        ):
            recommendation = "Buy - Good DLMM metrics"
        elif (
            pair["scores"]["total"] >= Config.HOLD_SCORE
            and pair["risk"]["rating"] <= Config.HIGH_RISK_RATING
        ):
            recommendation = "Hold - Monitor bin efficiency"
        else:
            recommendation = "Watch - Suboptimal DLMM setup"

        pair["recommendation"] = {
            "action": recommendation,
            "reason": f"Score: {pair['scores']['total']:.1f}/100, Risk: {pair['risk']['rating']}/5 - {pair['risk']['description']}",
        }

        return pair


class Logger:
    @staticmethod
    def setup():
        # Clear any existing handlers
        root = logging.getLogger()
        for handler in root.handlers[:]:
            root.removeHandler(handler)

        # Configure root logger
        logging.basicConfig(
            level=logging.INFO,
            format="%(message)s",
            stream=sys.stdout,
        )

        # Add file handler for detailed logging
        file_handler = logging.handlers.RotatingFileHandler(
            Config.LOG_FILE,
            maxBytes=1024 * 1024,
            backupCount=5,
        )
        file_handler.setFormatter(
            logging.Formatter("%(asctime)s - %(levelname)s - %(message)s"),
        )
        file_handler.setLevel(logging.DEBUG)
        logging.getLogger().addHandler(file_handler)

        return logging.getLogger(__name__)


async def get_jupiter_token_list(api_client: ApiClient) -> dict[str, dict[str, Any]]:
    """Fetch Jupiter token list from API."""
    try:
        data = await api_client.fetch(Config.JUPITER_TOKEN_STRICT_LIST_API)
        if data and isinstance(data, list):
            return {token["address"]: token for token in data}
    except Exception:
        logging.exception("Failed to fetch Jupiter token list")
        return {}
    else:
        return {}


async def get_meteora_pairs(api_client: ApiClient) -> list[dict[str, Any]]:
    """Fetch Meteora pairs and filter out inactive ones."""
    try:
        url = f"{Config.METEORA_API}{Config.METEORA_API_PAIR_ENDPOINT}"
        data = await api_client.fetch(url)
        if data and isinstance(data, list):
            # Pre-filter pairs to reduce unnecessary API calls
            active_pairs = []
            for pair in data:
                # Skip pairs with zero or very low liquidity
                if (
                    not pair.get("liquidity")
                    or float(pair.get("liquidity", 0)) < Config.MIN_LIQUIDITY_THRESHOLD
                ):
                    continue

                # Skip pairs with extreme bin steps (likely inactive or test pools)
                bin_step = int(pair.get("bin_step", 0))
                if bin_step <= 0 or bin_step > Config.MAX_BIN_STEP:
                    continue

                # Skip pairs with zero fees (but allow high fees if they have liquidity)
                base_fee = float(pair.get("base_fee_percentage", 0))
                if base_fee <= 0:  # Only skip zero fees
                    continue

                active_pairs.append(pair)

            logging.info(
                "Filtered %d Meteora pairs down to %d active pairs",
                len(data),
                len(active_pairs),
            )
            return active_pairs
    except Exception:
        logging.exception("Failed to fetch Meteora pairs")
        return []
    else:
        return []


def addresses_to_dex_screener_urls(addresses: list[str]) -> list[str]:
    """Convert addresses to DexScreener API URLs with proper batching."""
    fetch_urls = [Config.DEX_SCREENER_API_URL]
    address_count = 0

    for address in addresses:
        current_url_index = len(fetch_urls) - 1
        address_count += 1

        if len(fetch_urls[current_url_index]) == len(Config.DEX_SCREENER_API_URL):
            fetch_urls[current_url_index] = f"{fetch_urls[current_url_index]}/{address}"
        else:
            updated_url = f"{fetch_urls[current_url_index]},{address}"
            if address_count < Config.CHUNK_SIZE:
                fetch_urls[current_url_index] = updated_url
            else:
                fetch_urls.append(f"{Config.DEX_SCREENER_API_URL}/{address}")
                address_count = 0

    return fetch_urls


def process_dex_screener_response(response: Any) -> list[dict[str, Any]]:
    """Process a single DexScreener response and return valid pairs."""
    if isinstance(response, Exception) or not response:
        return []

    if not isinstance(response, dict) or "pairs" not in response:
        return []

    pairs = response["pairs"]
    if not isinstance(pairs, list):
        return []

    return [
        p
        for p in pairs
        if isinstance(p, dict)
        and "pairAddress" in p
        and "baseToken" in p
        and "quoteToken" in p
        and "volume" in p
        and "liquidity" in p
    ]


@dataclass
class ProgressParams:
    """Parameters for calculating progress metrics."""

    chunk_group_idx: int
    total_pairs: int
    total_addresses: int
    start_time: float
    concurrent_limit: int
    chunk_size: int
    address_chunks: list[list[str]]


def calculate_progress(params: ProgressParams) -> tuple[int, float, float, float]:
    """Calculate progress metrics for batch processing."""
    processed_addresses = min(
        params.chunk_group_idx * params.concurrent_limit * params.chunk_size,
        params.total_addresses,
    )
    progress = (processed_addresses / params.total_addresses) * 100
    elapsed_time = time.time() - params.start_time
    pairs_per_second = processed_addresses / elapsed_time if elapsed_time > 0 else 0
    remaining_addresses = params.total_addresses - processed_addresses
    eta_seconds = (
        (remaining_addresses / pairs_per_second)
        if pairs_per_second > 0
        else float("inf")
    )
    return processed_addresses, progress, pairs_per_second, eta_seconds


async def get_dex_screener_pairs(
    api_client: ApiClient,
    addresses: list[str],
) -> list[dict[str, Any]]:
    """Fetch DexScreener data using batched requests with parallel processing"""
    if not addresses:
        return []

    results = []
    total_pairs = 0
    total_addresses = len(addresses)
    start_time = time.time()
    processing_queue = asyncio.Queue()

    async def process_responses():
        while True:
            try:
                response = await processing_queue.get()
                if response is None:  # Sentinel value to stop processing
                    break

                valid_pairs = process_dex_screener_response(response)
                nonlocal total_pairs
                total_pairs += len(valid_pairs)
                results.extend(valid_pairs)
                processing_queue.task_done()
            except Exception:
                logging.exception("Error processing response")
                processing_queue.task_done()

    # Start processor task
    processor = asyncio.create_task(process_responses())

    # Create batches of addresses
    address_chunks = [
        addresses[i : i + Config.CHUNK_SIZE]
        for i in range(0, len(addresses), Config.CHUNK_SIZE)
    ]

    # Process multiple chunks concurrently while respecting rate limits
    concurrent_chunks = [
        address_chunks[i : i + Config.CONCURRENT_LIMIT]
        for i in range(0, len(address_chunks), Config.CONCURRENT_LIMIT)
    ]

    for chunk_group_idx, chunk_group in enumerate(concurrent_chunks):
        chunk_tasks = []

        for addresses_chunk in chunk_group:
            addresses_str = ",".join(addresses_chunk)
            url = f"{Config.DEX_SCREENER_API_URL}/{addresses_str}"
            task = api_client.fetch(url)
            chunk_tasks.append(task)

        # Process concurrent chunk group
        responses = await asyncio.gather(*chunk_tasks, return_exceptions=True)

        # Add responses to processing queue
        for response in responses:
            await processing_queue.put(response)

        # Wait for all responses in this group to be processed
        await processing_queue.join()

        # Update progress
        (
            processed_addresses,
            progress_percent,
            pairs_per_second,
            eta_seconds,
        ) = calculate_progress(
            ProgressParams(
                chunk_group_idx,
                total_pairs,
                total_addresses,
                start_time,
                Config.CONCURRENT_LIMIT,
                Config.CHUNK_SIZE,
                address_chunks,
            ),
        )

        logging.info(
            "Progress: %d/%d (%.1f%%) - Found %d pairs (%.1f pairs/sec, ETA: %.1f min)",
            processed_addresses,
            total_addresses,
            progress_percent,
            total_pairs,
            pairs_per_second,
            eta_seconds / 60,
        )

        # Small delay between chunk groups to respect rate limit
        await asyncio.sleep(Config.BASE_DELAY)

    # Signal processor to stop and wait for it to finish
    await processing_queue.put(None)
    await processor

    logging.info("Completed fetching %d pairs from DexScreener", total_pairs)
    return results


def create_hyperlink(url: str, text: str, for_excel: bool = False) -> str:
    """Create a hyperlink for CSV or Excel output."""
    if for_excel:
        return f"[{text}] {url}"
    return url


def save_to_csv(
    data: list[dict[str, Any]],
    filename: str = "market_opportunities.csv",
) -> None:
    """Save analyzed data to CSV file."""
    if not data:
        logging.warning("No data to save")
        return

    try:
        headers = [
            "Recommendation",
            "Score",
            "Risk Rating",
            "Pair",
            "Base",
            "Quote",
            "Bin Step",
            "Fees %",
            "TVL",
            "FDV",
            "5m Vol",
            "1h Vol",
            "24h Vol",
            "5m Fees",
            "1h Fees",
            "24h Fees",
            "5m Vol/TVL",
            "1h Vol/TVL",
            "24h Vol/TVL",
            "5m Momentum",
            "1h Momentum",
            "6h Momentum",
            "Score Details",
            "Risk Factors",
            "Strict List",
            "Base RugCheck",
            "Quote RugCheck",
            "DEX Screener",
            "Meteora Market",
        ]

        with Path(filename).open("w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(headers)

            # Sort by score and risk rating
            for pair in sorted(
                data,
                key=lambda x: (x["scores"]["total"], -x["risk"]["rating"]),
                reverse=True,
            ):
                # Format score components
                score_details = " | ".join(
                    f"{k}: {v:.1f}" for k, v in pair["scores"]["components"].items()
                )

                # Format risk factors
                risk_factors = " | ".join(
                    factor
                    for factor, present in pair["risk"]["factors"].items()
                    if present
                )

                row = [
                    pair["recommendation"]["action"],
                    f"{pair['scores']['total']:.1f}",
                    f"{pair['risk']['rating']} - {pair['risk']['description']}",
                    f"{pair['baseToken']['symbol']}-{pair['quoteToken']['symbol']}",
                    pair["baseToken"]["symbol"],
                    pair["quoteToken"]["symbol"],
                    pair["bin_step"],
                    f"{pair['base_fee'] * 100:.1f}%",
                    f"${pair['tvl']:,.2f}",
                    f"${pair['fdv']:,.2f}",
                    f"${pair['volume']['raw']['m5']:,.2f}",
                    f"${pair['volume']['raw']['h1']:,.2f}",
                    f"${pair['volume']['raw']['h24']:,.2f}",
                    f"${pair['fees']['raw']['m5']:,.2f}",
                    f"${pair['fees']['raw']['h1']:,.2f}",
                    f"${pair['fees']['raw']['h24']:,.2f}",
                    f"{pair['ratios']['volume_to_tvl']['m5']:.3f}",
                    f"{pair['ratios']['volume_to_tvl']['h1']:.3f}",
                    f"{pair['ratios']['volume_to_tvl']['h24']:.3f}",
                    f"{pair['momentum']['5m']:.2f}",
                    f"{pair['momentum']['1h']:.2f}",
                    f"{pair['momentum']['6h']:.2f}",
                    score_details,
                    risk_factors if risk_factors else "None",
                    "Yes" if pair["strict"] else "No",
                    create_hyperlink(
                        f"https://rugcheck.xyz/tokens/{pair['baseToken']['address']}",
                        f"{pair['baseToken']['symbol']} RugCheck",
                    )
                    if pair["baseToken"]["symbol"].lower() not in Config.BLUE_CHIPS
                    else "Blue Chip",
                    create_hyperlink(
                        f"https://rugcheck.xyz/tokens/{pair['quoteToken']['address']}",
                        f"{pair['quoteToken']['symbol']} RugCheck",
                    )
                    if pair["quoteToken"]["symbol"].lower() not in Config.BLUE_CHIPS
                    else "Blue Chip",
                    create_hyperlink(
                        f"https://dexscreener.com/solana/{pair['pairAddress']}",
                        "Price Chart",
                    ),
                    create_hyperlink(
                        f"https://app.meteora.ag/dlmm/{pair['pairAddress']}",
                        "Trade Now",
                    ),
                ]
                writer.writerow(row)

        logging.info("Successfully saved %d rows to %s", len(data), filename)
    except Exception:
        logging.exception("Failed to save CSV file")
        raise


def apply_volume_formatting(cell: Any, value: str) -> None:
    """Apply blue highlighting for volume columns."""
    try:
        if float(str(value).replace("$", "").replace(",", "")) > 0:
            cell.fill = PatternFill(
                start_color="E3F2FD",
                end_color="E3F2FD",
                fill_type="solid",
            )
    except ValueError:
        pass


def apply_fee_formatting(cell: Any, value: str) -> None:
    """Apply purple highlighting for fee columns."""
    try:
        if float(str(value).replace("$", "").replace(",", "")) > 0:
            cell.fill = PatternFill(
                start_color="F3E5F5",
                end_color="F3E5F5",
                fill_type="solid",
            )
    except ValueError:
        pass


def apply_momentum_formatting(cell: Any, value: str) -> None:
    """Apply green/red highlighting for momentum columns."""
    try:
        value_float = float(str(value))
        if value_float > 0:
            cell.fill = PatternFill(
                start_color="E8F5E9" if value_float > 1 else "FFEBEE",
                end_color="E8F5E9" if value_float > 1 else "FFEBEE",
                fill_type="solid",
            )
    except ValueError:
        pass


def adjust_column_width(ws: Any, column: Any) -> None:
    """Adjust column width based on content."""
    max_length = 0
    col = column[0].column_letter
    for cell in column:
        with suppress(Exception):
            max_length = max(len(str(cell.value)), max_length)
    adjusted_width = max_length + 2
    ws.column_dimensions[col].width = min(adjusted_width, 50)


def create_worksheet() -> tuple[openpyxl.Workbook, Worksheet]:
    """Create and validate Excel worksheet."""
    wb = openpyxl.Workbook()
    active_sheet = wb.active
    if active_sheet is None:
        msg = "Failed to create worksheet"
        logging.error(msg)
        raise ValueError(msg)
    active_sheet.title = "Market Opportunities"
    return wb, active_sheet


def save_to_excel(
    data: list[dict[str, Any]],
    filename: str = "market_opportunities.xlsx",
) -> None:
    """Save analyzed data to Excel file with DexScreener-style formatting."""
    if not data:
        logging.warning("No data to save")
        return

    try:
        wb, ws = create_worksheet()

        # Write headers
        headers = [
            "Recommendation",
            "Score",
            "Risk Rating",
            "Pair",
            "Base",
            "Quote",
            "Bin Step",
            "Fees %",
            "TVL",
            "FDV",
            "5m Vol",
            "1h Vol",
            "24h Vol",
            "5m Fees",
            "1h Fees",
            "24h Fees",
            "5m Vol/TVL",
            "1h Vol/TVL",
            "24h Vol/TVL",
            "5m Momentum",
            "1h Momentum",
            "6h Momentum",
            "Score Details",
            "Risk Factors",
            "Strict List",
            "Base RugCheck",
            "Quote RugCheck",
            "DEX Screener",
            "Meteora Market",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)

        # Sort data by score and risk rating
        sorted_data = sorted(
            data,
            key=lambda x: (x["scores"]["total"], -x["risk"]["rating"]),
            reverse=True,
        )

        # Write data rows
        for row_idx, pair in enumerate(sorted_data, 2):
            score_details = " | ".join(
                f"{k}: {v:.1f}" for k, v in pair["scores"]["components"].items()
            )
            risk_factors = " | ".join(
                factor for factor, present in pair["risk"]["factors"].items() if present
            )

            row = [
                pair["recommendation"]["action"],
                f"{pair['scores']['total']:.1f}",
                f"{pair['risk']['rating']} - {pair['risk']['description']}",
                f"{pair['baseToken']['symbol']}-{pair['quoteToken']['symbol']}",
                pair["baseToken"]["symbol"],
                pair["quoteToken"]["symbol"],
                pair["bin_step"],
                f"{pair['base_fee'] * 100:.1f}%",
                f"${pair['tvl']:,.2f}",
                f"${pair['fdv']:,.2f}",
                f"${pair['volume']['raw']['m5']:,.2f}",
                f"${pair['volume']['raw']['h1']:,.2f}",
                f"${pair['volume']['raw']['h24']:,.2f}",
                f"${pair['fees']['raw']['m5']:,.2f}",
                f"${pair['fees']['raw']['h1']:,.2f}",
                f"${pair['fees']['raw']['h24']:,.2f}",
                f"{pair['ratios']['volume_to_tvl']['m5']:.3f}",
                f"{pair['ratios']['volume_to_tvl']['h1']:.3f}",
                f"{pair['ratios']['volume_to_tvl']['h24']:.3f}",
                f"{pair['momentum']['5m']:.2f}",
                f"{pair['momentum']['1h']:.2f}",
                f"{pair['momentum']['6h']:.2f}",
                score_details,
                risk_factors if risk_factors else "None",
                "Yes" if pair["strict"] else "No",
                create_hyperlink(
                    f"https://rugcheck.xyz/tokens/{pair['baseToken']['address']}",
                    f"{pair['baseToken']['symbol']} RugCheck",
                    True,
                )
                if pair["baseToken"]["symbol"].lower() not in {*Config.BLUE_CHIPS}
                else "Blue Chip",
                create_hyperlink(
                    f"https://rugcheck.xyz/tokens/{pair['quoteToken']['address']}",
                    f"{pair['quoteToken']['symbol']} RugCheck",
                    True,
                )
                if pair["quoteToken"]["symbol"].lower() not in {*Config.BLUE_CHIPS}
                else "Blue Chip",
                create_hyperlink(
                    f"https://dexscreener.com/solana/{pair['pairAddress']}",
                    "Price Chart",
                    True,
                ),
                create_hyperlink(
                    f"https://app.meteora.ag/dlmm/{pair['pairAddress']}",
                    "Trade Now",
                    True,
                ),
            ]

            for col, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.value = value

                # Apply hyperlink formatting for link columns
                if col in [26, 27, 28, 29]:  # Link columns
                    cell.font = Font(color="0000FF", underline="single")

                # Apply color formatting
                if col in [11, 12, 13]:  # Volume columns
                    apply_volume_formatting(cell, value)
                elif col in [14, 15, 16]:  # Fee columns
                    apply_fee_formatting(cell, value)
                elif col in [20, 21, 22]:  # Momentum columns
                    apply_momentum_formatting(cell, value)

        # Create table
        tab = openpyxl.worksheet.table.Table(
            displayName="MarketOpportunities",
            ref=f"A1:{openpyxl.utils.get_column_letter(len(headers))}{len(sorted_data) + 1}",
            tableStyleInfo=openpyxl.worksheet.table.TableStyleInfo(
                name="TableStyleMedium12",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            ),
        )
        ws.add_table(tab)

        # Freeze header row
        ws.freeze_panes = ws["A2"]

        # Auto-adjust column widths
        for column in ws.columns:
            adjust_column_width(ws, column)

        wb.save(filename)
        logging.info("Successfully saved %d rows to %s", len(data), filename)
    except Exception:
        logging.exception("Failed to save Excel file")
        raise


def save_data(
    data: list[dict[str, Any]],
    output_format: str = "csv",
) -> None:
    """Save data in the specified format."""
    if output_format == "xlsx":
        save_to_excel(data)
    else:
        save_to_csv(data)


async def main():
    # Parse command line arguments
    parser = argparse.ArgumentParser(description="Meteora DLMM Pool Analyzer")
    parser.add_argument(
        "--format",
        choices=["csv", "xlsx"],
        default="csv",
        help="Output format (default: csv)",
    )
    args = parser.parse_args()

    logger = Logger.setup()
    logger.info("Starting Meteora analysis")

    try:
        # Initialize API client
        api_client = ApiClient()

        # Step 1: Fetch token list
        logger.info("Fetching Jupiter token list...")
        token_map = await get_jupiter_token_list(api_client)
        logger.info("Retrieved %d tokens from Jupiter", len(token_map))

        # Step 2: Fetch Meteora pairs
        logger.info("Fetching Meteora pairs...")
        meteora_pairs = await get_meteora_pairs(api_client)
        logger.info("Retrieved %d Meteora pairs", len(meteora_pairs))
        addresses = [str(pair["address"]) for pair in meteora_pairs]

        # Step 3: Fetch DexScreener data
        logger.info("Fetching DexScreener data...")
        dex_screener_pairs = await get_dex_screener_pairs(api_client, addresses)

        # Step 4: Process and analyze data
        logger.info("Analyzing opportunities...")
        opportunities = []
        for pair in dex_screener_pairs:
            enriched_pair = DataProcessor.enrich_pair_data(
                pair,
                token_map,
                meteora_pairs,
            )
            if enriched_pair:
                opportunities.append(enriched_pair)

        # Step 5: Save results
        logger.info("Saving results...")
        save_data(opportunities, args.format)

        logger.info("Analysis completed successfully")
    except Exception:
        logger.exception("Analysis failed")
        raise
    finally:
        await api_client.session.close()


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        logging.info("Analysis interrupted by user")
        sys.exit(0)
    except Exception:
        logging.exception("Fatal error")
        sys.exit(1)
